#!/usr/bin/env python3
"""
Animon Fakemon Sync Script

Coleta dados de:
- DeviantArt (imagens + descrições) via scraping das páginas públicas
- Planilha Excel (campos estruturados como fallback)

Gera: data/animons.json com 741 animon completos
"""

import json
import re
import time
import openpyxl
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path
from html.parser import HTMLParser

# ---------------------------------------------------------------------------
# CONFIGURAÇÃO
# ---------------------------------------------------------------------------

ALBUMS = {
    'Brasca':    ('https://www.deviantart.com/metanikk/gallery/57375612/brasca-region-first-generation', True),
    'Dunnottar': ('https://www.deviantart.com/metanikk/gallery/62148368/dunnottar-region-second-generation', True),
    'Lekker':    ('https://www.deviantart.com/metanikk/gallery/64370949/lekker-region-third-generation', True),
    'Rebueno':   ('https://www.deviantart.com/metanikk/gallery/70862073/rebueno-region-fourth-region', True),
    'Unknown':   ('https://www.deviantart.com/metanikk/gallery/57136001/unknown-dex', False),
    'Solar':     ('https://www.deviantart.com/metanikk/gallery/75262884/system-solar-dex', True),
    'Turtle':    ('https://www.deviantart.com/metanikk/gallery/69802096/turtle-dex', True),
    'Paleta':    ('https://www.deviantart.com/metanikk/gallery/78589649/paleta-mexicana-dex', True),
}

TYPE_MAP = {
    'Normal':'Neutral','Fire':'Fire','Fighting':'Combat','Water':'Water','Flying':'Flying',
    'Grass':'Botanical','Poison':'Venom','Electric':'Electric','Ground':'Chthonic',
    'Psychic':'Mentis','Rock':'Mineral','Ice':'Frost','Bug':'Arthropod','Dragon':'Dragon',
    'Ghost':'Spectral','Dark':'Shadow','Steel':'Metal','Fairy':'Mythic',
    'Cosmic':'Cosmic','Sound':'Sound','Tribal':'Tribal','Fungal':'Fungal',
    'Machine':'Machine','Microbial':'Microbial','Neutral':'Neutral'
}

TYPE_EMOJI = {
    'Botanical':'🌿','Water':'💧','Shadow':'🌑','Arthropod':'🦗','Spectral':'👻',
    'Fire':'🔥','Mentis':'🔮','Neutral':'⭐','Combat':'👊','Dragon':'🐉',
    'Electric':'⚡','Mineral':'🪨','Chthonic':'🌍','Venom':'☠️','Frost':'❄️',
    'Flying':'🦅','Metal':'⚙️','Mythic':'🌸','Cosmic':'✦','Sound':'🔊',
    'Tribal':'🥁','Fungal':'🍄','Machine':'🤖','Microbial':'🦠'
}

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
}

# ---------------------------------------------------------------------------
# HTTP HELPERS
# ---------------------------------------------------------------------------

def fetch_url(url, retries=3):
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=15) as resp:
                return resp.read().decode('utf-8', errors='replace')
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
            else:
                print(f'  ✗ Erro ao buscar {url}: {e}')
                return None

# ---------------------------------------------------------------------------
# DEVIANTART SCRAPING
# ---------------------------------------------------------------------------

def extract_deviation_urls_from_gallery(gallery_url, skip_first=True):
    """Extrai URLs de deviations de uma página de galeria do DeviantArt."""
    html = fetch_url(gallery_url)
    if not html:
        return []

    # Busca links de deviations no HTML
    # Formato: https://www.deviantart.com/metanikk/art/TITULO-123456
    pattern = r'href="(https://www\.deviantart\.com/metanikk/art/[^"]+)"'
    urls = list(dict.fromkeys(re.findall(pattern, html)))  # dedup mantendo ordem

    # Também tenta extrair do JSON embutido (DeviantArt usa hydration)
    json_pattern = r'"url"\s*:\s*"(https://www\.deviantart\.com/metanikk/art/[^"]+)"'
    urls2 = list(dict.fromkeys(re.findall(json_pattern, html)))
    
    all_urls = list(dict.fromkeys(urls + urls2))

    if skip_first and all_urls:
        all_urls = all_urls[1:]  # Ignora banner

    return all_urls


def extract_number_from_title(title):
    """Extrai número regional do título da deviation. Ex: '001 - Myrmeflag' → 1"""
    m = re.match(r'^(\d+)\s*[-–—]', title.strip())
    if m:
        return int(m.group(1))
    # Tenta número no final ou meio
    m = re.search(r'\b(\d{1,3})\b', title)
    if m:
        return int(m.group(1))
    return None


def parse_description(text):
    """Extrai campos estruturados da descrição de uma deviation."""
    result = {}
    if not text:
        return result

    # Campos simples linha a linha
    field_patterns = {
        'name':        r'(?:^|\n)\s*Name\s*:\s*(.+)',
        'type_raw':    r'(?:^|\n)\s*Type\s*:\s*(.+)',
        'ability':     r'(?:^|\n)\s*Ability\s*:\s*(.+)',
        'ha':          r'(?:^|\n)\s*HA\s*:\s*(.+)',
        'species':     r'(?:^|\n)\s*Species\s*:\s*(.+)',
        'height':      r'(?:^|\n)\s*Height\s*:\s*(.+)',
        'weight':      r'(?:^|\n)\s*Weight\s*:\s*(.+)',
        'habitat':     r'(?:^|\n)\s*Habitat\s*:\s*(.+)',
        'feeding':     r'(?:^|\n)\s*Feeding\s*:\s*(.+)',
        'evolution':   r'(?:^|\n)\s*Evolution\s*:\s*(.+)',
        'gender':      r'(?:^|\n)\s*Gender\s*(?:Ratio)?\s*:\s*(.+)',
    }
    for field, pat in field_patterns.items():
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            if val.lower() not in ('does not have', 'none', '-', ''):
                result[field] = val

    # Descrição PT-BR
    m = re.search(r'Descri[çc][aã]o\s*:\s*(.+?)(?=\nDescription|\nInspira|\nInspiration|\Z)', text, re.IGNORECASE | re.DOTALL)
    if m:
        result['description_pt'] = m.group(1).strip()

    # Descrição EN
    m = re.search(r'\nDescription\s*:\s*(.+?)(?=\nDescri|\nInspira|\nInspiration|\Z)', text, re.IGNORECASE | re.DOTALL)
    if m:
        result['description_en'] = m.group(1).strip()

    return result


def scrape_deviation(url):
    """Scrapa uma deviation individual: título, imagem, descrição."""
    html = fetch_url(url)
    if not html:
        return None

    result = {'url': url, 'title': None, 'image_url': None, 'description_raw': None}

    # Título — og:title ou <title>
    m = re.search(r'<meta\s+property="og:title"\s+content="([^"]+)"', html)
    if m:
        result['title'] = m.group(1).replace(' by Metanikk on DeviantArt', '').strip()
    else:
        m = re.search(r'<title>([^<]+)</title>', html)
        if m:
            result['title'] = m.group(1).split(' by ')[0].strip()

    # Imagem — og:image (maior resolução disponível)
    m = re.search(r'<meta\s+property="og:image"\s+content="([^"]+)"', html)
    if m:
        img = m.group(1)
        # Tenta pegar versão maior removendo parâmetros de resize
        img = re.sub(r'/v1/fill/[^/]+/', '/v1/fill/w_1200,h_1200,q_100,strp/', img)
        result['image_url'] = img

    # Descrição — procura no JSON de hydration (campo "description" ou "body")
    # DeviantArt injeta dados como __NEXT_DATA__ ou similar
    m = re.search(r'"description"\s*:\s*"((?:[^"\\]|\\.)*)"', html)
    if m:
        raw = m.group(1)
        raw = raw.encode('utf-8').decode('unicode_escape', errors='replace')
        # Remove tags HTML
        raw = re.sub(r'<[^>]+>', '\n', raw)
        raw = re.sub(r'&amp;', '&', raw)
        raw = re.sub(r'&lt;', '<', raw)
        raw = re.sub(r'&gt;', '>', raw)
        raw = re.sub(r'&#\d+;', '', raw)
        raw = re.sub(r'\n{3,}', '\n\n', raw).strip()
        result['description_raw'] = raw

    return result

# ---------------------------------------------------------------------------
# EXCEL LOADER
# ---------------------------------------------------------------------------

def type_enum(t):
    if not t:
        return None
    return TYPE_MAP.get(str(t).strip(), str(t).strip())


def load_excel():
    excel_path = Path(__file__).parent.parent / 'Fakemon Region.xlsx'
    if not excel_path.exists():
        print(f'⚠ Excel não encontrado: {excel_path}')
        return {}

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    animons_by_key = {}

    def safe(r, i):
        return r[i] if i < len(r) and r[i] is not None else None

    def insp(*cols):
        return ' / '.join([str(c) for c in cols if c]) or None

    # BRASCA
    ws = wb['Brasca']
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if not safe(r,1) or str(safe(r,1)).strip() in ('','NAME'): continue
        num = int(r[0])
        if num > 200: break
        animons_by_key[f'Brasca_{num}'] = {
            'regional': num, 'mr': num, 'name': str(r[1]).strip(),
            'type1': type_enum(safe(r,2)), 'type2': type_enum(safe(r,3)),
            'species': str(r[4]).strip() if safe(r,4) else None,
            'ability1': str(r[5]).strip() if safe(r,5) else None,
            'ability2': str(r[6]).strip() if safe(r,6) else None,
            'ha': None,
            'height': str(r[7]) if safe(r,7) else None,
            'weight': str(r[8]) if safe(r,8) else None,
            'habitat': None, 'feeding': None, 'evo_by': None,
            'inspiration': insp(*[safe(r,i) for i in range(9,15)])
        }

    # DUNNOTTAR
    ws = wb['Dunnottar']
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if not safe(r,1) or str(safe(r,1)).strip() in ('','NAME'): continue
        reg = int(r[0]); mr = 200 + reg
        animons_by_key[f'Dunnottar_{reg}'] = {
            'regional': reg, 'mr': mr, 'name': str(r[1]).strip(),
            'type1': type_enum(safe(r,2)), 'type2': type_enum(safe(r,3)),
            'species': str(r[4]).strip() if safe(r,4) else None,
            'ability1': str(r[5]).strip() if safe(r,5) else None,
            'ability2': str(r[6]).strip() if safe(r,6) else None,
            'ha': str(r[7]).strip() if safe(r,7) else None,
            'height': str(r[8]) if safe(r,8) else None,
            'weight': str(r[9]) if safe(r,9) else None,
            'habitat': str(r[10]) if safe(r,10) else None,
            'feeding': str(r[11]) if safe(r,11) else None,
            'evo_by': str(r[12]) if safe(r,12) and str(r[12])!='-' else None,
            'inspiration': insp(safe(r,13), safe(r,14))
        }

    # LEKKER
    ws = wb['Lekker']
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if not safe(r,1) or str(safe(r,1)).strip() in ('','NAME'): continue
        reg = int(r[0]); mr = 300 + reg
        animons_by_key[f'Lekker_{reg}'] = {
            'regional': reg, 'mr': mr, 'name': str(r[1]).strip(),
            'type1': type_enum(safe(r,2)), 'type2': type_enum(safe(r,3)),
            'species': str(r[4]).strip() if safe(r,4) else None,
            'ability1': str(r[5]).strip() if safe(r,5) else None,
            'ability2': str(r[6]).strip() if safe(r,6) else None,
            'ha': str(r[7]).strip() if safe(r,7) else None,
            'height': str(r[8]) if safe(r,8) else None,
            'weight': str(r[9]) if safe(r,9) else None,
            'habitat': str(r[10]) if safe(r,10) else None,
            'feeding': str(r[11]) if safe(r,11) else None,
            'evo_by': str(r[12]) if safe(r,12) and str(r[12])!='-' else None,
            'inspiration': insp(safe(r,13), safe(r,14))
        }

    # REBUENO
    ws = wb['Rebueno']
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if not safe(r,2) or str(safe(r,2)).strip() in ('','NAME'): continue
        mr = int(r[0]); reg = int(r[1])
        animons_by_key[f'Rebueno_{reg}'] = {
            'regional': reg, 'mr': mr, 'name': str(r[2]).strip(),
            'type1': type_enum(safe(r,3)), 'type2': type_enum(safe(r,4)),
            'species': str(r[5]).strip() if safe(r,5) else None,
            'ability1': str(r[6]).strip() if safe(r,6) else None,
            'ability2': str(r[7]).strip() if safe(r,7) else None,
            'ha': str(r[8]).strip() if safe(r,8) else None,
            'height': str(r[9]) if safe(r,9) else None,
            'weight': str(r[10]) if safe(r,10) else None,
            'habitat': str(r[11]) if safe(r,11) else None,
            'feeding': str(r[12]) if safe(r,12) else None,
            'evo_by': str(r[13]) if safe(r,13) and str(r[13])!='-' else None,
            'inspiration': insp(safe(r,14), safe(r,15))
        }

    # UNKNOWN
    ws = wb['Unknown']
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if not safe(r,1) or str(safe(r,1)).strip() in ('','NAME'): continue
        mr = int(r[0])
        animons_by_key[f'Unknown_{mr}'] = {
            'regional': mr, 'mr': mr, 'name': str(r[1]).strip(),
            'type1': type_enum(safe(r,2)), 'type2': type_enum(safe(r,3)),
            'species': str(r[4]).strip() if safe(r,4) else None,
            'ability1': str(r[5]).strip() if safe(r,5) else None,
            'ability2': str(r[6]).strip() if safe(r,6) else None,
            'ha': str(r[7]).strip() if safe(r,7) else None,
            'height': str(r[8]) if safe(r,8) else None,
            'weight': str(r[9]) if safe(r,9) else None,
            'habitat': str(r[10]) if safe(r,10) else None,
            'feeding': str(r[11]) if safe(r,11) else None,
            'evo_by': str(r[12]) if safe(r,12) and str(r[12])!='-' else None,
            'inspiration': insp(safe(r,13), safe(r,14), safe(r,15))
        }

    # MINIDEX
    def mini_region(n):
        if n <= 513: return 'Zodiac'
        if n <= 522: return 'Solar'
        if n <= 540: return 'Turtle'
        return 'Paleta'

    ws = wb['MiniDex']
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if not safe(r,1) or str(safe(r,1)).strip() in ('','NAME'): continue
        mr = int(r[0]); region = mini_region(mr)
        animons_by_key[f'{region}_{mr}'] = {
            'regional': mr, 'mr': mr, 'name': str(r[1]).strip(),
            'type1': type_enum(safe(r,2)), 'type2': type_enum(safe(r,3)),
            'species': str(r[4]).strip() if safe(r,4) else None,
            'ability1': str(r[5]).strip() if safe(r,5) else None,
            'ability2': str(r[6]).strip() if safe(r,6) else None,
            'ha': str(r[7]).strip() if safe(r,7) else None,
            'height': str(r[8]) if safe(r,8) else None,
            'weight': str(r[9]) if safe(r,9) else None,
            'habitat': str(r[10]) if safe(r,10) else None,
            'feeding': str(r[11]) if safe(r,11) else None,
            'evo_by': str(r[12]) if safe(r,12) and str(r[12])!='-' else None,
            'inspiration': insp(safe(r,13), safe(r,14), safe(r,15))
        }

    wb.close()
    return animons_by_key

# ---------------------------------------------------------------------------
# DEVIANTART ENRICHMENT
# ---------------------------------------------------------------------------

def build_name_index(animons_by_key):
    """Índice nome → key para matching por nome."""
    idx = {}
    for key, data in animons_by_key.items():
        name = data.get('name', '').lower().strip()
        if name:
            idx[name] = key
    return idx


def fetch_all_deviations(animons_by_key):
    """Para cada álbum, scrapa deviations e faz matching com o banco."""
    name_idx = build_name_index(animons_by_key)
    enriched = {}  # key → {image_url, description_en, description_pt, deviantart_url}

    for region, (album_url, skip_banner) in ALBUMS.items():
        print(f'\n📂 Buscando álbum: {region}')
        deviation_urls = extract_deviation_urls_from_gallery(album_url, skip_first=skip_banner)
        print(f'   {len(deviation_urls)} deviations encontradas')

        for i, dev_url in enumerate(deviation_urls):
            time.sleep(0.6)  # Rate limiting
            dev = scrape_deviation(dev_url)
            if not dev or not dev.get('title'):
                continue

            title = dev['title']
            num = extract_number_from_title(title)

            # Tenta matching por número regional + região
            matched_key = None
            if num:
                # Monta key candidata
                if region == 'Unknown':
                    # Unknown usa MR direto (541-746 range)
                    candidate = f'Unknown_{num}'
                    if candidate in animons_by_key:
                        matched_key = candidate
                elif region in ('Solar', 'Turtle', 'Zodiac', 'Paleta'):
                    candidate = f'{region}_{num}'
                    if candidate in animons_by_key:
                        matched_key = candidate
                else:
                    candidate = f'{region}_{num}'
                    if candidate in animons_by_key:
                        matched_key = candidate

            # Fallback: matching por nome
            if not matched_key:
                # Extrai nome do título: "001 - Myrmeflag" → "myrmeflag"
                name_part = re.sub(r'^\d+\s*[-–—]\s*', '', title).lower().strip()
                if name_part in name_idx:
                    matched_key = name_idx[name_part]

            if not matched_key:
                print(f'   ⚠ Sem match: "{title}"')
                continue

            parsed = parse_description(dev.get('description_raw', ''))

            enriched[matched_key] = {
                'image_url': dev.get('image_url'),
                'description_en': parsed.get('description_en'),
                'description_pt': parsed.get('description_pt'),
                'deviantart_url': dev_url,
            }
            name = animons_by_key[matched_key].get('name', matched_key)
            print(f'   ✓ [{i+1}] {name} ← "{title}"')

    return enriched

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def generate_animons(use_deviantart=True):
    print('📋 Carregando planilha Excel...')
    excel_data = load_excel()
    print(f'   {len(excel_data)} animon carregados do Excel')

    enriched = {}
    if use_deviantart:
        print('\n🌐 Buscando imagens e descrições do DeviantArt...')
        try:
            enriched = fetch_all_deviations(excel_data)
            print(f'\n   ✓ {len(enriched)} animon enriquecidos com DeviantArt')
        except Exception as e:
            print(f'\n   ⚠ Erro no DeviantArt: {e}. Continuando sem imagens.')

    animons = []
    for key, data in excel_data.items():
        mr = data['mr']
        region = key.split('_')[0]
        t1 = data.get('type1')
        dev = enriched.get(key, {})

        animon = {
            'id': mr,
            'mr': mr,
            'regional': int(data.get('regional', mr)),
            'region': region,
            'name': data.get('name', f'Animon#{mr}'),
            'type1': t1,
            'type2': data.get('type2'),
            'species': data.get('species'),
            'ability1': data.get('ability1'),
            'ability2': data.get('ability2'),
            'ha': data.get('ha'),
            'height': data.get('height'),
            'weight': data.get('weight'),
            'habitat': data.get('habitat'),
            'feeding': data.get('feeding'),
            'evo_by': data.get('evo_by'),
            'inspiration': data.get('inspiration'),
            'emoji': TYPE_EMOJI.get(t1, '❔'),
            'image_url': dev.get('image_url'),
            'description_en': dev.get('description_en'),
            'description_pt': dev.get('description_pt'),
            'deviantart_url': dev.get('deviantart_url'),
            'last_updated': datetime.now().isoformat() + 'Z'
        }
        animons.append(animon)

    animons.sort(key=lambda x: x['mr'])
    return animons


if __name__ == '__main__':
    import sys
    use_da = '--no-deviantart' not in sys.argv

    animons = generate_animons(use_deviantart=use_da)

    data_dir = Path(__file__).parent.parent / 'data'
    data_dir.mkdir(exist_ok=True)

    data_path = data_dir / 'animons.json'
    with open(data_path, 'w', encoding='utf-8') as f:
        json.dump(animons, f, ensure_ascii=False, indent=2)

    with_images = sum(1 for a in animons if a.get('image_url'))
    metadata = {
        'total': len(animons),
        'with_images': with_images,
        'last_sync': datetime.now().isoformat() + 'Z',
        'regions': len(set(a['region'] for a in animons))
    }
    with open(data_dir / 'metadata.json', 'w', encoding='utf-8') as f:
        json.dump(metadata, f, indent=2)

    print(f'\n✅ {len(animons)} animon sincronizados')
    print(f'🖼  {with_images} com imagem do DeviantArt')
    print(f'📁 Arquivo: {data_path}')
